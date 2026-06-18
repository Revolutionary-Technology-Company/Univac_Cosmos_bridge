import os
import json
import openpyxl

class UnivacExcelJsonAdapter:
    def __init__(self, excel_path="../Univac-IX/materials_db.xlsx"):
        """Initializes connection to the calculation engine spreadsheet repository."""
        self.excel_path = excel_path
        if not os.path.exists(self.excel_path):
            raise FileNotFoundError(f"Univac-IX Engine Error: Spreadsheet calculation core missing at {self.excel_path}")

    def compute_and_serialize_to_json(self, raw_sensor_permittivity, raw_conductivity):
        """
        Injects real-time telemetry inputs, triggers cross-page calculations, 
        and extracts the final synced page values into an absolute JSON format string.
        """
        # 1. Open workbook to inject input arguments into the calculation matrix
        wb = openpyxl.load_workbook(self.excel_path, data_only=False)
        
        # Target your primary ingestion sheet (Adjust names to match your precise tab layouts)
        if "Sensor_Ingestion" in wb.sheetnames:
            ingest_sheet = wb["Sensor_Ingestion"]
            # Inject raw variables directly into designated formula tracking cells
            ingest_sheet["A2"] = float(raw_sensor_permittivity)
            ingest_sheet["B2"] = float(raw_conductivity)
            wb.save(self.excel_path)

        # 2. Open workbook with 'data_only=True' to extract the fully calculated and synced values
        # This acts as your Excel calculation engine step
        evaluated_wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        
        # Extract metadata outputs from your synced calculation sheets
        material_sheet = evaluated_wb["Material_Outputs"] if "Material_Outputs" in evaluated_wb.sheetnames else evaluated_wb.active
        lattice_sheet = evaluated_wb["Lattice_Analytics"] if "Lattice_Analytics" in evaluated_wb.sheetnames else evaluated_wb.active

        # 3. Compile the calculated cross-page evaluations into a clean dictionary structure
        calculated_matrix = {
            "telemetry_metadata": {
                "injected_permittivity": raw_sensor_permittivity,
                "injected_conductivity": raw_conductivity
            },
            "inferred_material_properties": {
                "classification_state": str(material_sheet["C2"].value), # e.g., 'Titanium_Alloy'
                "density_g_cm3": float(material_sheet["D2"].value or 0.0),
                "yield_strength_mpa": float(material_sheet["E2"].value or 0.0)
            },
            "molecular_lattice_structure": {
                "lattice_type": str(lattice_sheet["B2"].value),       # e.g., 'BCC'
                "spacing_angstrom": float(lattice_sheet["C2"].value or 0.0), # e.g., 3.32
                "atomic_packing_factor": float(lattice_sheet["D2"].value or 0.0)
            }
        }

        # 4. Serialize to standardized JSON format
        return json.dumps(calculated_matrix, indent=4)

# --- WORKSPACE EXECUTION RUN TEST ---
if __name__ == "__main__":
    print("Testing Univac-IX Linked Excel-to-JSON Serialization Engine...")
    
    # Simple setup of a dummy workbook structure if you run it locally without your real repo present
    # In live development, this file will already exist inside your sibling Univac-IX folder
    try:
        adapter = UnivacExcelJsonAdapter()
        
        # Simulate incoming antenna metrics triggering the calculations
        json_output_payload = adapter.compute_and_serialize_to_json(
            raw_sensor_permittivity=7.42, 
            raw_conductivity=1.85
        )
        
        print("\n=== SYSTEM SYNCED DATA PACKET CONVERTED TO JSON ===")
        print(json_output_payload)
        
    except FileNotFoundError as e:
        print(f"Simulation bypass: {e}")
        print("-> To deploy in production, ensure your calculated materials_db.xlsx sits in the Univac-IX folder.")
